import json
from pathlib import Path

from numpy import ndarray
import pandas as pd
import MDAnalysis as mda
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment

from ..logger import Logger
from ..data_converter import DataConverter


logger = Logger("FileWriter")


class FileWriter:
    @staticmethod
    def write_json_file(
            data: dict,
            path_to_file: Path,
    ) -> None:
        with Path(path_to_file).open("w") as json_file:
            json.dump(data, json_file, indent=4)

    @classmethod
    def write_dat_file(
            cls,
            data_lines: list[str] | ndarray,
            path_to_file: Path,
            to_overwrite: bool = True,
    ) -> None | Path:
        """
        Write .dat file.
        If data_lines is ndarray, it will be converted to list[str] using DataConverter.convert_ndarray_to_dat.
        """

        # TODO: refactor to use DataConverter.convert_ndarray_to_dat
        try:
            if len(data_lines) == 0:
                logger.warning("No data for .dat file.")
                return

            if to_overwrite is False and path_to_file.exists():
                # Don't to_overwrite existing file
                return

            if not path_to_file.suffix == ".dat":
                # Set .dat extension
                path_to_file = path_to_file.with_suffix(".dat")

            # Ensure the directory exists
            path_to_file.parent.mkdir(parents=True, exist_ok=True)

            # Convert ndarray to list[str]
            if isinstance(data_lines, ndarray):
                data_lines = DataConverter.convert_ndarray_to_dat(data_lines)

            # Write the data
            with Path(path_to_file).open("w") as dat_file:
                for line in data_lines:
                    if "\n" not in line:
                        line += "\n"
                    dat_file.write(line)

            logger.info(f"File saved: {path_to_file}")
            return path_to_file

        except Exception as e:
            logger.error(f".dat file not saved: {e}")

    @staticmethod
    def write_pdb_file(
            data_lines: list[str] | ndarray,
            path_to_file: Path,
            to_overwrite: bool = True,
    ) -> None | Path:
        """
        Write .pdb file.
        If data_lines is ndarray, it will be converted to list[str] using DataConverter.convert_ndarray_to_pdb.
        """

        try:
            if len(data_lines) == 0:
                logger.warning("No data for .dat file.")
                return

            if to_overwrite is False and path_to_file.exists():
                # Don't to_overwrite existing file
                return

            if not path_to_file.suffix == ".pdb":
                # Set .pdb extension
                path_to_file = path_to_file.with_suffix(".pdb")

            # Ensure the directory exists
            path_to_file.parent.mkdir(parents=True, exist_ok=True)

            # Convert ndarray to list[str]
            if isinstance(data_lines, ndarray):
                data_lines = DataConverter.convert_ndarray_to_pdb(data_lines)

            # Write the data
            with Path(path_to_file).open("w") as pdb_file:
                for line in data_lines:
                    pdb_file.write(line)

            logger.info(f"File saved: {path_to_file}")
            return path_to_file

        except Exception as e:
            logger.error(f".pdb file not saved: {e}")

    @classmethod
    def write_excel_file(
            cls,
            df: pd.DataFrame,
            sheet_name: str,
            path_to_file: Path,
    ) -> Path | None:
        """
        Write a pandas DataFrame to an Excel file.

        Parameters:
        - df: pd.DataFrame, the data to write to the Excel file.
        - structure_dir: str, the name of the structure folder.
        - folder_path: Path | str | None, the base folder path. If None, uses the default logic from PathBuilder.
        - file_name: str, the Excel file name to write.
        - sheet_name: str, the name of the sheet where data will be written.
        - is_init_data_dir: bool | None: to build path to the specific dir. If it's False - builds path to result data.
        """

        if not path_to_file.suffix == ".xlsx":
            # Set .xlsx extension
            path_to_file = path_to_file.with_suffix(".xlsx")

        # Ensure the directory exists
        path_to_file.parent.mkdir(parents=True, exist_ok=True)

        try:
            if isinstance(df.columns, pd.MultiIndex):
                # Handle MultiIndex columns
                with pd.ExcelWriter(path_to_file, engine='openpyxl') as writer:
                    # Write the DataFrame to the Excel file
                    df.to_excel(writer, sheet_name=sheet_name, startrow=1, header=False, index=True)

                    # Access the workbook and the sheet
                    worksheet: Worksheet = writer.sheets[sheet_name]

                    # Write the MultiIndex header
                    for idx, col in enumerate(df.columns):
                        worksheet.cell(row=1, column=idx + 2, value=col[0])
                        worksheet.cell(row=2, column=idx + 2, value=col[1])

                    # Merge cells for the top-level headers
                    for col in df.columns.levels[0]:
                        col_indices: list[int] = [i for i, c in enumerate(df.columns) if c[0] == col]
                        if col_indices:
                            worksheet.merge_cells(
                                start_row=1, start_column=col_indices[0] + 2,
                                end_row=1, end_column=col_indices[-1] + 2,
                            )
                            # Center align the merged cells
                            worksheet.cell(row=1, column=col_indices[0] + 2).alignment = Alignment(horizontal='center')
            else:
                # Write the DataFrame to an Excel file
                df.to_excel(path_to_file, sheet_name=sheet_name, index=False, engine="openpyxl")

            logger.info(f"Data successfully written to {path_to_file}")
            return path_to_file

        except Exception as e:
            logger.error(f"Failed to write file {path_to_file}: {e}")

    @staticmethod
    def write_pdb_from_mda(output_pdb_file: Path, atoms) -> None:
        with mda.Writer(output_pdb_file, n_atoms=atoms.n_atoms) as PDB:
            PDB.write(atoms)
